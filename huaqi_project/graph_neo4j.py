import openpyxl
from py2neo import Graph, Node, NodeMatcher, Relationship, RelationshipMatcher


class gupiaoGraph:
    def __init__(self,local_path,username,password):
        self.username=username
        self.password=password
        self.g=Graph(local_path,auth=(self.username,self.password),name="neo4j")
        self.g.delete_all()
        self.g.begin()
    def create_graph(self,path_down,path_up):
        work_sheet_down= openpyxl.load_workbook(path_down).active
        row_1=work_sheet_down.iter_rows(1,1)
        dict_key=[]
        columns=[]
        for row in row_1:
            for i,cell in enumerate(row):
                if cell.value=="参控公司":
                    columns.append(i)
                    dict_key.append("name")
                if cell.value=="持股比例(%)":
                    columns.append(i)
                    dict_key.append(cell.value)
                if cell.value=="注册资本(万元)":
                    columns.append(i)
                    dict_key.append(cell.value)
                if cell.value=="注册地":
                    columns.append(i)
                    dict_key.append(cell.value)
                if cell.value=="报告期":
                   columns.append(i)
                   dict_key.append(cell.value)
                if cell.value=="主营业务":
                   columns.append(i)
                   dict_key.append(cell.value)
                if cell.value=="证券简称":
                   column_company=i+1
                if cell.value=="参控关系":
                   column_relation=i
                   columns.append(i)
                   dict_key.append(cell.value)
        row_all=work_sheet_down.iter_rows(2)
        node_sum=Node("母公司",**{"name":work_sheet_down.cell(2,column_company).value})
        self.g.create(node_sum)
        for row in row_all:
           dict_value=[]
           for i in columns:
               dict_value.append(row[i].value)
           dict_node=dict(zip(dict_key,dict_value))
           node_son=Node("子公司",**dict_node)
           self.g.create(node_son)
           relation=Relationship(node_sum,row[column_relation].value,node_son)
           self.g.create(relation)
        work_sheet_up= openpyxl.load_workbook(path_up).active
        row_1=work_sheet_up.iter_rows(1,1)
        columns_down=[]
        dic_key=[]
        for row in row_1:
            for i,cell in enumerate(row):
                if cell.value=="股东名称":
                    columns_down.append(i)
                    dic_key.append("name")
                if cell.value=="持股数量":
                    columns_down.append(i)
                    dic_key.append(cell.value)
                if cell.value=="持股比例":
                    columns_down.append(i)
                    column_relation=i
                    dic_key.append(cell.value)
                if cell.value=="持股变化":
                    columns_down.append(i)
                    dic_key.append(cell.value)
                if cell.value=="报告类型":
                    columns_down.append(i)
                    dic_key.append(cell.value)
                if cell.value=="公告日期":
                    columns_down.append(i)
                    dic_key.append(cell.value)
        row_all=work_sheet_up.iter_rows(2)
        for row in row_all:
            dic_value=[]
            for i in columns_down:
                dic_value.append(row[i].value)
            dic_node=dict(zip(dic_key,dic_value))
            if len(dic_node["name"])>4:
                node_up=Node("持股公司",**dic_node)
            else:
                node_up=Node("持股人",**dic_node)
            self.g.create(node_up)
            relation=Relationship(node_up,"股东",node_sum,**{"持股比例":row[column_relation].value})
            self.g.create(relation)
    def get_entity(self,proportion):
        node_matcher=NodeMatcher(self.g)
        nodes=node_matcher.match("子公司")
        node_need=[]
        for node in nodes:
            if node["持股比例(%)"]>proportion:
                node_need.append(node)
        for node in node_need:
             print(node["name"])
    def get_class_entity(self,class_name):
        node_matcher=NodeMatcher(self.g)
        nodes=node_matcher.match("子公司")
        node_need=[]
        for node in nodes:
            if node["参控关系"]==class_name:
                print(node["name"])
            if node["主营业务"]==class_name:
                print(node["name"])
    def get_gudong_person(self):
        node_matcher=NodeMatcher(self.g)
        nodes=node_matcher.match("持股人")
        for node in nodes:
            print(node["name"])
    def get_gudong_company(self):
        node_matcher=NodeMatcher(self.g)
        nodes=node_matcher.match("持股公司")
        for node in nodes:
            print(node["name"])
        

